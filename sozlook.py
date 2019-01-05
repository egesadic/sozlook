import os
import sys
import requests
from bs4 import BeautifulSoup
import pickle 
import re
import time
import random
import deturkizer
from openpyxl import Workbook
from datetime import datetime
import json


#Global variables for sozlook to use.
PATH = os.getcwd()
SOZLUKLER = {"eksi": "eksisozluk.com/"}
USER_AGENT = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36'}
SELECTED_SOZLUK = None
TOPIC = None
HTML = None
LINK = "http://"

class Entry(object):
    def __init__(self, author = "", aid = 0,  date = "", likes = 0, text = "", eid = 0):
        self.author = author
        self.aid = int(aid)
        self.date = date
        self.likes = int(likes)
        self.text = text
        self.eid = int(eid)

class ConsoleMessage(object):
    def __init__(self, message = ""):
        self.message = message

    def change_message(self, string):
        self.message = string

    def write(self):
        sys.stdout.write(self.message+"\r")
        sys.stdout.flush()    

    def remove_message(self):
        i = 0
        text = ""
        while i < len(self.message):
            text = text + " "
            i += 1
        sys.stdout.write(text + "\r")
        sys.stdout.flush()

def choose_sozluk(sozluk):
    global LINK 
    global SELECTED

    if SOZLUKLER.__contains__(sozluk):
        SELECTED = sozluk
        LINK = LINK + SOZLUKLER[sozluk]
        print ("Sözlük " + LINK + " olarak seçildi.")
    else:
        print ("Sözlük bulunamadı.")

def get_url(url):
    tmp = requests.get(url)
    html = BeautifulSoup(tmp.text, "html.parser")
    return html

def get_topic(topic, page_iter = False):
    global LINK
    global HTML
    global TOPIC

    if not page_iter:
        TOPIC = topic
    link = LINK + topic
    HTML = get_url(link)

def extract_date(entry):
    date = re.findall(r'\d{2}.\d{2}.\d{4}', entry)
    return date[0]

def format_entry(entry):
    text = entry
    text = text.replace("\n", "").replace("\r", "").replace("  ", "")
    date = re.findall(r'\d{2}.\d{2}.\d{4}', entry)
    text = text.replace(date[0], "")
    time = re.findall(r'\d{2}:\d{2}', entry)
    text = text.replace(time[0], "")
    return text

def format_query(string):
    q = str(string).replace(" ", "+")
    return q

def search(query, kw_search = False, t_list = [], page_count = 0, iter_mode = False):
    topic_list = t_list
    if not iter_mode:
        url = "https://www.google.com/search?q=" + format_query(query) + "+site:eksisozluk.com"
    else:
        url = query
    pages = page_count
    results = []
    cm = ConsoleMessage("")

    response = requests.get(url, headers=USER_AGENT)
    response.raise_for_status()
    html = BeautifulSoup(response.text, "html.parser")
    if kw_search:
        i = 0
        navend = html.find("a", id = "pnnext")
        l_next = navend['href']
        while l_next is not None:
            cm.change_message("Arama sonuçları indiriliyor, toplam indirilen sayfa: " + str(pages))
            cm.write()
            tag_cite = html.find_all("cite", class_ = "iUh30")
            tag_span = html.find_all("span", class_ = "st")
            while i < len(tag_cite):
                result = (tag_cite[i].text, tag_span[i].text)
                topic_list.append(result)
                i += 1
            del result
            cm.change_message("Sleeping to avoid Papa Google. Page" + str(pages))
            cm.remove_message()
            cm.write()
            time.sleep(15)
            pages += 1
            search("http://google.com" + str(l_next), kw_search, topic_list, pages, True)
            cm.remove_message()
        for elem in topic_list:
            tmp = requests.get(elem[0])
            html = BeautifulSoup(tmp.text, "html.parser")
            parsed = parse_entries(html)
            for item in parsed:
                if item.text == elem[1]:
                    results.append(item)
        return results

def parse_entries(html = None):
    global HTML
    if html is None:
        html = HTML
    entries = []

    print("Şimdi alınıyor: " + '"'+(html.title.text).replace(" - ekşi sözlük", "")+'"')
    ul = html.find("ul", id = "entry-item-list")
    if ul is not None:
        for li in ul.find_all("li"):
            entry = Entry(li["data-author"], li["data-author-id"], extract_date(li.text), li["data-favorite-count"], format_entry(li.text).replace(li["data-author"]), li["data-id"])
            entries.append(entry)
            del entry
    return entries

def parse_all_entries(html = None):
    global HTML
    global SELECTED
    html = HTML
    entries = []
    suffix = "?p="

    if SOZLUKLER.__contains__(SELECTED):
        pages = html.find_all("div", class_ = "pager")
        page_count = pages[0]["data-pagecount"]
        del pages
        i = 1
        while i <= int(page_count):
            entries.extend(parse_entries())
            i += 1
            tmp = TOPIC+suffix+str(i)
            get_topic(tmp, True)
            html = HTML
        return entries
    else:
        print ("Entry bulunamadı, lütfen başlığı kontrol ediniz.")

def load_entries(entrybase):
    pfile = open(entrybase, 'rb')
    entries = pickle.load(pfile)
    pfile.close()
    return entries

def save_entries(entries, toText = False):
    if isinstance(entries, list):
        title = (HTML.title.text).replace(" - ekşi sözlük", "")
        entryFile = open(title + "-entrybase", "wb")
        if toText:
            textFile = open(title + "-text-outout", "w")
            for elem in entries:
                textFile.write("Yazar: " + elem.author + "\nEntry (" + str(elem.likes) +" Beğeni) : " + elem.text)
            print ("Entryler, metin dosyasına kaydedildi.")
        entryFile = open(title + "-entrybase", "wb")
        pickle.dump(entries, entryFile)
        entryFile.close()
        print ("Entry dosyası oluşturuldu.")

def to_excel(entries, name):
    r = 2
    c = 1
    excel = Workbook()
    excel.create_sheet("Entries")
    sheet = excel["Entries"]
    sheet.cell(row = 1, column = 1).value = "Yazar"
    sheet.cell(row = 1, column = 2).value = "Tarih"
    sheet.cell(row = 1, column = 3).value = "Like"
    sheet.cell(row = 1, column = 4).value = "Entry"
    for item in entries:
        sheet.cell(row = r, column = c).value = item.author
        c += 1
        sheet.cell(row = r, column = c).value = item.date
        c += 1
        sheet.cell(row = r, column = c).value = item.likes
        c += 1
        sheet.cell(row = r, column = c).value = item.text
        r += 1
        c = 1
    excel.save(name + '.xlsx')
    print ("Excel dosyasına döküldü.")

def to_json(entries, name):
    with open(name + '.json', 'w') as f:
        f.write(json.dumps([ob.__dict__ for ob in entries], ensure_ascii= False))
        f.close()
       

